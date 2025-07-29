from click import ParamType

import csv

import openpyxl

from pathlib import Path

from logging import getLogger

import simplejson

import functools

from tabulate import tabulate

from datonius import *
import datonius.ontology as ontology

TABLE = functools.partial(tabulate, headers='keys', tablefmt='simple')
TSV = functools.partial(tabulate, headers='keys', tablefmt='tsv')
JSON = functools.partial(simplejson.dumps, indent=2, iterable_as_array=True)
OBJECTS = lambda o: o

def first_value_or_default(iterable, default=None):
    try:
        return next(iter(iterable))
    except StopIteration:
        class DefaultEmptyObject:
            def __init__(self, default):
                self.default = default

            def __getattr__(self, name):
                return self.default
        return DefaultEmptyObject(default)


def isolate_to_dict(name, iso, can_nest=False):
    if not iso:
        # click.echo(f"{name} not found.", err=True)
        return {}
    rec = dict(
            accession=iso.fda_accession,
            organism=str(iso.binomial),
            gims_barcode=iso.barcode,
            strain=iso.strain_name,
            pulsenet_key=iso.fda_pulse_net_key,
            biosample=iso.biosample,
            collection_date=iso.sample.collection_date.isoformat(),
            related_food=iso.sample.related_food
    )
    if can_nest:
        # formats where we can nest dictionaries
        rec['foodon'] = iso.sample.related_food_ontology_code
        rec['collection_date_range'] = f"{iso.sample.collection_date - iso.sample.collection_date_accuracy_mask} - {iso.sample.collection_date + iso.sample.collection_date_accuracy_mask}"
        rec['collection_reason'] = iso.sample.collection_reason
        rec['firms'] = [
            dict(
                rel_code = rel.relationship_code,
                firm_fei = rel.firm.fei,
                firm_name = rel.firm.name,
                firm_address = rel.firm.address.full,
                was_responsible = rel.responsibility
            )
        for rel in iso.sample.firm_relationships]
        #other stuff
    else:
        # alternate summary views
        rec['responsible_firm_fei'] = first_value_or_default(iso.sample.responsible_firms).fei
        rec['responsible_firm_name'] = first_value_or_default(iso.sample.responsible_firms).name
    return rec

def sample_to_dict(sam):
    return dict(

    )


def lookup(names, output_format=OBJECTS, all_fields=False):
    use_full_data = all_fields or output_format is JSON
    return output_format([isolate_to_dict(name, Isolate.of(name), use_full_data) for name in names])

def tax(names, output_format=OBJECTS, all_fields=False):
    use_full_data = all_fields or output_format is JSON
    return output_format((isolate_to_dict(names[-1], iso, use_full_data) for iso in Taxon.get(name=names[-1]).isolates))

def ontology(names, output_format=OBJECTS, all_fields=True):
    use_full_data = all_fields or output_format is JSON
    return output_format(isolate_to_dict())

# Better home for a lot of the general utility in the datonius-etl package



log = getLogger('datonius.util')

class AnnotatedGenerator:

    def __init__(self, name, generator):
        self.name = name
        self.generator = generator

    def __iter__(self):
        yield from self.generator
    
    def __str__(self):
        return f"<row-like generator from file '{self.name}'>"

    def __repr__(self):
        return self.__str__()

class TabularFile(ParamType):
    name = "csv, tsv, txt, or xlsx"

    def get_row_reader(self, file):
        "Open file with correct reader"
        log.debug(f"got {file}")
        if hasattr(file, 'name'): # is an open file handle whose name we can read
            log.debug(f"{file} is an fp")
            if any(map(lambda ext: ext in file.name, ('csv', 'tsv', 'txt'))):
                log.debug(f"tabular file extension")
                yield from csv.reader(file, dialect='excel', delimiter=',' if 'csv' in file.name else '\t')
            else:
                log.debug(f"Possible Excel file")
                wb = openpyxl.load_workbook(file, read_only=True)
                sheet = wb.active
                yield from filter(lambda row: row, (tuple(cell.value or '' for cell in row) for row in sheet.iter_rows()) )
        else:
            log.debug(f"String or path")
            if any(map(lambda ext: ext in file, ('csv', 'tsv', 'txt'))):
                log.debug(f"tabular file extension")
                with open(file, 'rt') as fi:
                    yield from csv.reader(fi, dialect='excel', delimiter=',' if 'csv' in file else '\t')
            else:
                log.debug(f"Possible Excel file")
                wb = openpyxl.load_workbook(filename=file, read_only=True)
                sheet = wb.active
                yield from filter(lambda row: row, (tuple(cell.value or ''  for cell in row) for row in sheet.iter_rows()) )


    def convert(self, value, param, ctx):
        "Open file with correct reader and return an iterator over its rows"
        return AnnotatedGenerator(value, self.get_row_reader(value))
    
class HeadedTabularFile(TabularFile):

    def convert(self, value, param, ctx):
        "Open file with correct reader then return a dict-like reader over rows"
        row_iter = self.get_row_reader(value)
        headers = tuple(value for value in next(row_iter))
        return AnnotatedGenerator(value, filter(lambda row: row, ({header:val for header, val in zip(headers, row) if header} for row in row_iter)))
