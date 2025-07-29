#!/usr/bin/env python

"""Tests for various tabular readers."""

import csv
from pathlib import Path

import unittest
from tempfile import NamedTemporaryFile as temp

from contextlib import contextmanager

import datonius.util as d

from openpyxl import Workbook

import io

from logging import getLogger

log = getLogger('test_readers')


class TestFileReaders(unittest.TestCase):
    """Tests for dict-like file readers"""

    def setUp(self):
        self.fr = d.TabularFile()

        self.data = [
            ['a', 'b', 'c'],
            ['1', '2', '3'],
            ['4', '5', '6'],
            ['7', '', '9'],
            ['', 'squirrel', '']
        ]

        self.dicts = tuple({key:val for key,val in zip(self.data[0], row)} for row in self.data[1:])

        self.file = None

    @contextmanager
    def tempfile(self, suffix, mode='w'):
        log.debug(f"File ending character on this platform is {repr(csv.get_dialect('excel').lineterminator)}")
        with temp(suffix=suffix, mode=mode, delete=False, newline='') as file:
            wtr = csv.writer(file, dialect='excel', delimiter=',' if 'csv' in suffix else '\t')
            wtr.writerows(self.data)
            self.file = Path(file.name)
        with open(self.file, 'rt', newline=csv.get_dialect('excel').lineterminator) as file:
            yield file
        # yield file.name
            

    def tearDown(self):
        if self.file:
            log.debug(f"destructing file {self.file}, was:")
            with open(self.file, 'rb') as file:
                log.debug(repr(file.read()))
            self.file.unlink()

    def test_whether_something_is_wrong_with_files(self):
        f = io.StringIO()
        wtr = csv.writer(f, dialect='excel', lineterminator='\n')
        wtr.writerows(self.data)
        f.seek(0)
        self.assertEqual(f.read(), "\n".join([",".join(row) for row in self.data]) + '\n')

    def test_whether_something_is_wrong_with_rows(self):
        f = io.StringIO("\n".join([",".join(row) for row in self.data]) + '\n')
        f.seek(0)
        def y():
            yield from csv.reader(f)
        self.assertListEqual(self.data, list(y()))


    def test_reading_writing(self):
        with self.tempfile('.csv') as file:
            rdr = csv.reader(file, dialect='excel')
            data = list(rdr)
            self.assertListEqual(self.data, data)

    def test_file_is_what_I_think_it_is(self):
        with self.tempfile('.csv') as file:
            L = file.readline()
            self.assertEqual(L, ",".join(self.data[0]) + csv.get_dialect('excel').lineterminator)

    def test_iterator(self):
        with self.tempfile('.csv') as file:
            L = list(self.fr.get_row_reader(file))
            self.assertEqual(len(L), 5)

    def test_first_line(self):
        with self.tempfile('.csv') as file:
            L = list(self.fr.get_row_reader(file))
            self.assertListEqual(L[0], ["a", "b", "c"])


    def test_implementation(self):
        with self.tempfile('.csv') as file:
            def y():
                yield from csv.reader(file, dialect='excel', delimiter=',' if 'csv' in file.name else '\t')
            L = list(y())
            file.seek(0)
            rdr = csv.reader(file, dialect='excel', delimiter=',')
            self.assertListEqual(list(rdr), L)

    @unittest.skip("")
    def test_yield_from_isnt_empty(self):
        f = io.StringIO("\n".join([",".join(row) for row in self.data]) + '\n')
        f.seek(0)
        f.name = '.csv'
        L = list(self.fr.get_row_reader(f))
        self.assertNotEqual(L, [])

    def test_why_yield_from_is_empty(self):
        L = [1, 2, 3]
        def y():
            yield from L
        self.assertEqual(list(L), list(y()))


    #@unittest.skip('Windows file pointers have unexpected behavior')
    def test_csv_reader(self):
        with self.tempfile(suffix='.csv') as file:
            data = list(self.fr.convert(file, None, None))
            self.assertListEqual(self.data, data)

    #@unittest.skip('Windows file pointers have unexpected behavior')
    def test_tsv_reader(self):
        with self.tempfile(suffix='.tsv') as file:
            data = list(self.fr.convert(file, None, None))
            self.assertListEqual(self.data, data)

    #@unittest.skip('Windows file pointers have unexpected behavior')
    def test_txt_reader(self):
        with self.tempfile(suffix='.txt') as file:
            data = list(self.fr.convert(file, None, None))
            self.assertListEqual(self.data, data)

    @unittest.skip("openpyxl doesn't support XLS")
    def test_xls_reader(self):
        with temp(suffix='.xls', mode='w+', encoding='utf-8', newline='') as file:
            wb = Workbook()
            ws = wb.active
            for x, row in enumerate(self.data):
                for y, value in enumerate(row):
                    ws.cell(row=x+1, column=y+1, value=value)
            wb.save(file)
            file.seek(0)
            data = tuple(self.fr.convert(file, None, None))
            self.assertTupleEqual(self.dicts, data)


class TestDictReaders(unittest.TestCase):

    tempfile = TestFileReaders.tempfile


    def setUp(self):
        TestFileReaders.setUp(self)
        self.fr = d.HeadedTabularFile()

    def tearDown(self):
        TestFileReaders.tearDown(self)

    

    def test_csv_reader(self):
        with self.tempfile(suffix='.csv') as file:
            data = tuple(self.fr.convert(file, None, None))
            self.assertTupleEqual(self.dicts, data)

    #@unittest.skip('Windows file pointers have unexpected behavior')
    def test_tsv_reader(self):
        with self.tempfile(suffix='.tsv') as file:
            data = tuple(self.fr.convert(file, None, None))
            self.assertTupleEqual(self.dicts, data)

    #@unittest.skip('Windows file pointers have unexpected behavior')
    def test_txt_reader(self):
        with self.tempfile(suffix='.txt') as file:
            data = tuple(self.fr.convert(file, None, None))
            self.assertTupleEqual(self.dicts, data)


class TestExcelReaders(unittest.TestCase):


    def setUp(self):
        TestFileReaders.setUp(self)

    #@unittest.skip('Windows file pointers have unexpected behavior')
    def test_xlsx_reader(self):
        with temp(suffix='.xlsx', mode='wb', delete=False) as file:
            pass
        self.file = file
        wb = Workbook()
        ws = wb.active
        for x, row in enumerate(self.data):
            for y, value in enumerate(row):
                ws.cell(row=x+1, column=y+1, value=value)
        wb.save(file.name)
        data = list(self.fr.convert(file.name, None, None))
        for a, b in zip(self.data, data):
            self.assertSequenceEqual(a, b)