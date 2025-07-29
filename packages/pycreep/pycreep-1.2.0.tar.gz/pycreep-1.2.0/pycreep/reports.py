"""
    Module to write pretty reports given some analysis results
"""

import numpy as np

from openpyxl import Workbook

from pycreep import allowables


def write_complete_report(
    fname, rupture_model, rate_model, yield_model, temperatures, times, confidence=0.9
):
    """
    Write something like a complete report out to an Excel file

    Args:
        fname:          file name to write to
        rupture_model:  trained creep rupture model
        rate_model:     trained creep rate model
        temperatures:   temperature increments to use in printing allowables
        times:          time increments to use in printing allowables

    Kwargs:
        confidence:     what confidence interval you want in getting
                        minimum information
    """
    wb = Workbook()

    # Write a tab with some  yield stress data
    tab = wb.create_sheet("Yield stress model")
    tab["A1"] = "Tabulated yield strengths"
    preds = yield_model(temperatures)
    tab["A2"] = "Temperature"
    tab["B2"] = "Value"
    for i, (T, p) in enumerate(zip(temperatures, preds)):
        tab.cell(row=i + 3, column=1, value=T)
        tab.cell(row=i + 3, column=2, value=p)

    # Write the rupture and rate models
    rupture_model.write_excel_report(wb, tabname="Rupture")
    rate_model.write_excel_report(wb, tabname="Strain rate")

    res = allowables.Sc_SectionII_1A_1B(
        rupture_model, rate_model, temperatures, full_results=True
    )

    # Write a tab with the allowable stress information
    tab = wb.create_sheet("Allowable stress")
    tab["A1"] = "Allowable stress information"
    headers = ["Temperature", "Savg", "n", "Favg", "FSavg", "Smin", "FSmin", "Sc", "S"]
    for i, h in enumerate(headers):
        tab.cell(row=2, column=1 + i, value=h)
    for j, T in enumerate(temperatures):
        tab.cell(row=3 + j, column=1, value=T)
        for i, h in enumerate(headers[1:]):
            tab.cell(row=3 + j, column=2 + i, value=res[h][j])

    # Write a tab with minimum stress to rupture information
    tab = wb.create_sheet("Minimum stress to rupture")
    tab["A1"] = "Minimum stress to rupture"
    tab["B1"] = "Time..."
    tab["A2"] = "Temperature"
    for i, t in enumerate(times):
        tab.cell(row=2, column=2 + i, value=t)
    for j, T in enumerate(temperatures):
        tab.cell(row=3 + j, column=1, value=T)
        for i, t in enumerate(times):
            tab.cell(
                row=3 + j,
                column=2 + i,
                value=rupture_model.predict_stress(
                    np.array([t]), np.array([T]), confidence=confidence
                )[0],
            )

    # Get rid of stupid default sheet
    del wb[wb.sheetnames[0]]

    wb.save(fname)
