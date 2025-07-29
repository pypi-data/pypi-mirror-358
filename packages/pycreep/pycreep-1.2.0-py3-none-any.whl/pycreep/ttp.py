"""Correlate time dependent data with a time-temperature parameter"""

import abc

import numpy as np
import scipy.stats
from openpyxl import Workbook

from pycreep import methods, dataset


class TTPAnalysis(dataset.DataSet):
    """
    Superclass for time-temperature parameter (TTP) analysis of a
    dataset

    Args:
        data:                       dataset as a pandas dataframe

    Keyword Args:
        time_field (str):           field in array giving time, default is
                                    "Life (h)"
        temp_field (str):           field in array giving temperature, default
                                    is "Temp (C)"
        stress_field (str):         field in array giving stress, default is
                                    "Stress (MPa)"
        heat_field (str):           filed in array giving heat ID, default is
                                    "Heat/Lot ID"
        input_temp_units (str):     temperature units, default is "C"
        input_stress_units (str):   stress units, default is "MPa"
        input_time_units (str):     time units, default is "hr"
        analysis_temp_units (str):  temperature units for analysis,
                                    default is "K"
        analysis_stress_units (str):    analysis stress units, default is
                                        "MPa"
        analysis_time_units (str):  analysis time units, default is "hr"
        time_sign (float):          sign to apply to time units, typically 1.0
                                    but for some analysis -1 makes sense

    The setup and analyzed objects are suppose to maintain the following properties:
        * "preds":      predictions for each point
        * "C_avg":      overall TTP parameter
        * "C_heat":     dictionary mapping each heat to the
                        lot-specific TTP
        * "poly_avg":   polynomial coefficients for the average
                        model
        * "R2":         coefficient of determination
        * "SSE":        standard squared error
        * "SEE":        standard error estimate
        * "SEE_heat":   SEE without lot centering, i.e. if you have a random heat
        * "R2_heat":    R2 without lot centering, i.e. if you have a random heat
    """

    def __init__(
        self,
        data,
        time_field="Life (h)",
        temp_field="Temp (C)",
        stress_field="Stress (MPa)",
        heat_field="Heat/Lot ID",
        input_temp_units="degC",
        input_stress_units="MPa",
        input_time_units="hrs",
        analysis_temp_units="K",
        analysis_stress_units="MPa",
        analysis_time_units="hrs",
        time_sign=1.0,
    ):
        super().__init__(data)

        self.add_field_units(
            "temperature", temp_field, input_temp_units, analysis_temp_units
        )
        self.add_field_units(
            "stress", stress_field, input_stress_units, analysis_stress_units
        )
        self.add_field_units("time", time_field, input_time_units, analysis_time_units)

        self.analysis_time_units = analysis_time_units
        self.analysis_temp_units = analysis_temp_units
        self.analysis_stress_units = analysis_stress_units

        self.add_heat_field(heat_field)
        self.time_sign = time_sign

    def excel_report(self, fname, tabname="Rupture"):
        """
        Write out an excel workbook

        Args:
            fname:      filename to use

        Kwargs:
            tabname:    what tab name to use
        """
        wb = Workbook()

        self.write_excel_report(wb, tabname=tabname)

        # Get rid of the dumb default tab
        del wb[wb.sheetnames[0]]

        wb.save(fname)

    def write_excel_report(self, wb, tabname="Rupture"):
        """
        Write to a particular excel workbook

        Args:
            wb:         workbook object

        Kwargs:
            tabname:    what tab name to use

        """
        tab = wb.create_sheet(tabname)

        self.write_excel_report_to_tab(tab)


class PolynomialAnalysis(TTPAnalysis):
    """
    Superclass for polynomial TTP analysis

    Args:
        input_ttp:                  time-temperature parameter
        order:                      polynomial order
        data:                       dataset as a pandas dataframe

    Keyword Args:
        time_field (str):           field in array giving time, default is
                                    "Life (h)"
        temp_field (str):           field in array giving temperature, default
                                    is "Temp (C)"
        stress_field (str):         field in array giving stress, default is
                                    "Stress (MPa)"
        heat_field (str):           filed in array giving heat ID, default is
                                    "Heat/Lot ID"
        input_temp_units (str):     temperature units, default is "C"
        input_stress_units (str):   stress units, default is "MPa"
        input_time_units (str):     time units, default is "hr"
        analysis_temp_units (str):  temperature units for analysis,
                                    default is "K"
        analysis_stress_units (str):    analysis stress units, default is
                                        "MPa"
        analysis_time_units (str):  analysis time units, default is "hr"

    The setup and analyzed objects are suppose to maintain the following properties:
        * "preds":      predictions for each point
        * "C_avg":      overall TTP parameter
        * "C_heat":     dictionary mapping each heat to the
                        lot-specific TTP
        * "poly_avg":   polynomial coefficients for the average
                        model
        * "R2":         coefficient of determination
        * "SSE":        standard squared error
        * "SEE":        standard error estimate
        * "SEE_heat":   SEE without lot centering, i.e. if you have a random heat
        * "R2_heat":    R2 without lot centering, i.e. if you have a random heat
    """

    def __init__(self, input_ttp, order, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.TTP = input_ttp
        self.order = order

    def report(self):
        """
        Provide a standard dict description of results
        """
        return {
            "preds": self.preds,
            "C_avg": self.C_avg,
            "C_heat": self.C_heat,
            "polyavg": self.polyavg,
            "R2": self.R2,
            "SSE": self.SSE,
            "SEE": self.SEE,
            "SEE_heat": self.SEE_heat,
            "R2_heat": self.R2_heat,
            "heat_count": {h: len(i) for h, i in self.heat_indices.items()},
            "heat_rms": self.heat_rms,
        }

    def write_excel_report_to_tab(self, tab):
        """
        Write an excel report to a given tab

        Args:
            tab (openpyxl tab):     tab handle to write to
        """
        tab["A1"] = "Regression results:"
        tab["A2"] = "Coefficient"
        tab["B2"] = "Value"
        of = 3
        for i, p in enumerate(self.polyavg[::-1]):
            tab.cell(row=i + 3, column=1, value=f"a{i}")
            tab.cell(row=i + 3, column=2, value=p)
        of = 3 + len(self.polyavg)
        tab.cell(row=of, column=1, value="Overall C:")
        tab.cell(row=of, column=2, value=self.C_avg)
        of += 2

        tab.cell(row=of, column=1, value="Statistics:")
        of += 1
        tab.cell(row=of, column=1, value="R2")
        tab.cell(row=of, column=2, value=self.R2_heat)
        of += 1
        tab.cell(row=of, column=1, value="SEE")
        tab.cell(row=of, column=2, value=self.SEE_heat)
        of += 2

        tab.cell(row=of, column=1, value="Heat summary:")
        of += 1
        tab.cell(row=of, column=1, value="Heat")
        tab.cell(row=of, column=2, value="Count")
        tab.cell(row=of, column=3, value="Lot C")
        tab.cell(row=of, column=4, value="Lot RMS error")
        of += 1

        heat_count = {h: len(i) for h, i in self.heat_indices.items()}

        for heat in sorted(self.C_heat.keys()):
            tab.cell(row=of, column=1, value=heat)
            tab.cell(row=of, column=2, value=heat_count[heat])
            tab.cell(row=of, column=3, value=self.C_heat[heat])
            tab.cell(row=of, column=4, value=self.heat_rms[heat])
            of += 1

    def __call__(self, stress, temperature):
        """
        Alias for self.predict_time(stress, temperature)

        Args:
            stress:         input stress values
            temperature:    input temperature values
        """
        return self.predict_time(stress, temperature)

    def predict_time(self, stress, temperature, confidence=None):
        """
        Predict new times given stress and temperature

        Args:
            stress:         input stress values
            temperature:    input temperature values

        Keyword Args:
            confidence:     confidence interval, if None provide
                            average predictions
        """
        if confidence is None:
            h = 0.0
        else:
            h = np.sign(confidence) * scipy.stats.norm.interval(np.abs(confidence))[1]

        return 10.0 ** (
            self.time_sign
            * self.TTP.predict(
                self.polyavg, self.C_avg + h * self.SEE_heat, stress, temperature
            )
        )

    def predict_stress(
        self, time, temperature, confidence=None, root_bounds=None, raise_on_error=True
    ):
        """
        Predict new values of stress given time and temperature

        Args:
            time:           input time values
            temperature:    input temperature values

        Keyword Args:
            confidence:     confidence interval, if None provide
                            average predictions
            root_bounds:    if not None, lower and upper bounds on which root value to use
                            when inverting the TTP polynomial
            raise_on_error: if true throw an error if we can't invert.  if not return nan
        """
        # Will want these sorted
        if root_bounds is not None:
            root_bounds = np.log10(np.sort(root_bounds))

        # Take the log of time
        ltime = self.time_sign * np.log10(time)

        if confidence is None:
            h = 0.0
        else:
            h = np.sign(confidence) * scipy.stats.norm.interval(np.abs(confidence))[1]

        # Calculate the TTP
        vals = self.TTP.value(
            self.C_avg + h * self.SEE_heat, time, temperature, time_sign=self.time_sign
        )

        def solve_one(x):
            pi = np.copy(self.polyavg)
            pi[-1] -= x
            rs = np.array(np.roots(pi))
            if np.all(np.abs(np.imag(rs)) > 0):
                if raise_on_error:
                    raise ValueError("Inverting relation to predict stress failed")
                return np.nan
            rs[np.abs(np.imag(rs)) > 0] = 0
            rs = np.real(rs)
            # Need to consider this...
            if root_bounds is None:
                return np.max(rs)
            val = np.logical_and(rs >= root_bounds[0], rs <= root_bounds[1])
            if np.all(np.logical_not(val)):
                if raise_on_error:
                    raise ValueError("No root falls within user provided bounds!")
                return np.nan
            return rs[val][0]

        # Solve each one, one at a time, for now
        # Vectorizing the cases with an analytic solution should be
        # possible
        if np.isscalar(vals):
            res = solve_one(vals)
        else:
            res = np.zeros_like(ltime)
            for i in range(len(ltime)):
                res[i] = solve_one(vals[i])

        return 10.0**res

    def predict_stress_discontinuous(self, time, temperature, confidence=None):
        """
        Predict new values of stress given time and temperature
        in an accurate but discontinuous way

        Args:
            time:           input time values
            temperature:    input temperature values

        Keyword Args:
            confidence:     confidence interval, if None provide
                            average predictions
        """
        return self.predict_stress(time, temperature, confidence=confidence)


class SplitAnalysis(TTPAnalysis):
    """
    Split the data into two halves based on some stress measure

    Args:
        stress_measure:             a TimeIndependentCorrelation providing
                                    the stress criteria
        fraction:                   the threshold is stress_measure(T) * fraction
        lower_model:                model for stress < threshold
        upper_mode:                 model for stress >= threshold
        data:                       dataset as a pandas dataframe

    Keyword Args:
        time_field (str):           field in array giving time, default is
                                    "Life (h)"
        temp_field (str):           field in array giving temperature, default
                                    is "Temp (C)"
        stress_field (str):         field in array giving stress, default is
                                    "Stress (MPa)"
        heat_field (str):           filed in array giving heat ID, default is
                                    "Heat/Lot ID"
        input_temp_units (str):     temperature units, default is "C"
        input_stress_units (str):   stress units, default is "MPa"
        input_time_units (str):     time units, default is "hr"
        analysis_temp_units (str):  temperature units for analysis,
                                    default is "K"
        analysis_stress_units (str):    analysis stress units, default is
                                        "MPa"
        analysis_time_units (str):  analysis time units, default is "hr"

    """

    def __init__(
        self, stress_measure, fraction, lower_model, upper_model, *args, **kwargs
    ):
        super().__init__(*args, **kwargs)

        self.stress_measure = stress_measure
        self.fraction = fraction

        self.lower_model = lower_model
        self.upper_model = upper_model

    def write_excel_report(self, wb, tabname="Rupture"):
        """
        Write to a particular excel workbook

        Args:
            wb:         workbook object

        Kwargs:
            tabname:    what tab name to use

        """
        tab = wb.create_sheet(tabname + ", upper stress range")
        self.upper_model.write_excel_report_to_tab(tab)

        tab = wb.create_sheet(tabname + ", lower stress range")
        self.lower_model.write_excel_report_to_tab(tab)

    def analyze(self):
        """
        Run/re-run analysis
        """
        self.stress_measure.analyze()

        # May be a better way to do this?
        thresh = self.threshold(self.temperature)
        self.lower_model.data = self.data[self.stress < thresh].reset_index(drop=True)
        self.lower_model.analyze()

        self.upper_model.data = self.data[self.stress >= thresh].reset_index(drop=True)
        self.upper_model.analyze()

        return self

    def threshold(self, temperature):
        """
        The threshold for switching between the two models

        Args:
            temperature:    input temperatures
        """
        return self.fraction * self.stress_measure.predict(temperature)

    def predict_time(self, stress, temperature, confidence=None):
        """
        Predict new times given stress and temperature

        Args:
            stress:         input stress values
            temperature:    input temperature values

        Keyword Args:
            confidence:     requested confidence interval
        """
        time = np.zeros_like(stress)

        thresh = self.threshold(temperature)

        time[stress < thresh] = self.lower_model.predict_time(
            stress, temperature, confidence
        )[stress < thresh]
        time[stress >= thresh] = self.upper_model.predict_time(
            stress, temperature, confidence
        )[stress >= thresh]

        return time

    def predict_stress(self, time, temperature, confidence=None):
        """
        Predict new values of stress given time and temperature
        in a smooth way

        Args:
            time:           input time values
            temperature:    input temperature values

        Keyword Args:
            confidence:     confidence interval, if None provide
                            average predictions
        """
        # Do the whole thing twice...
        upper = self.upper_model.predict_stress(time, temperature, confidence)
        lower = self.lower_model.predict_stress(time, temperature, confidence)

        return np.minimum(upper, lower)

    def predict_stress_discontinuous(self, time, temperature, confidence=None):
        """
        Predict new values of stress given time and temperature
        in an accurate but discontinuous way

        Args:
            time:           input time values
            temperature:    input temperature values

        Keyword Args:
            confidence:     confidence interval, if None provide
                            average predictions
        """
        # Do the whole thing twice...
        upper_none = self.upper_model.predict_stress(time, temperature)
        lower_none = self.lower_model.predict_stress(time, temperature)
        upper = self.upper_model.predict_stress(time, temperature, confidence)
        lower = self.lower_model.predict_stress(time, temperature, confidence)

        thresh = self.threshold(temperature)

        res = np.zeros_like(upper_none)
        res[upper_none >= thresh] = upper[upper_none >= thresh]
        res[lower_none < thresh] = lower[lower_none < thresh]

        neither = np.logical_and(upper_none < thresh, lower_none >= thresh)

        wf = (upper_none[neither] - thresh) / (
            upper_none[neither] - lower_none[neither]
        )

        res[neither] = (1 - wf) * upper[neither] + wf * lower[neither]

        return res


class UncenteredAnalysis(PolynomialAnalysis):
    """
    Do an uncentered analysis of the data

    Args:
        TTP:                        time-temperature parameter
        order:                      polynomial order
        data:                       dataset as a pandas dataframe

    Keyword Args:
        time_field (str):           field in array giving time, default is
                                    "Life (h)"
        temp_field (str):           field in array giving temperature, default
                                    is "Temp (C)"
        stress_field (str):         field in array giving stress, default is
                                    "Stress (MPa)"
        heat_field (str):           filed in array giving heat ID, default is
                                    "Heat/Lot ID"
        input_temp_units (str):     temperature units, default is "C"
        input_stress_units (str):   stress units, default is "MPa"
        input_time_units (str):     time units, default is "hr"
        analysis_temp_units (str):  temperature units for analysis,
                                    default is "K"
        analysis_stress_units (str):    analysis stress units, default is
                                        "MPa"
        analysis_time_units (str):  analysis time units, default is "hr"
    """

    def analyze(self):
        """
        Actually do the regression analysis and set standard properties
        """
        X = np.concatenate(
            (
                np.vander(np.log10(self.stress), N=self.order + 1)
                * self.TTP.stress_transform(self.time, self.temperature)[:, None],
                -np.ones((len(self.stress), 1)),
            ),
            axis=1,
        )
        y = self.time_sign * np.log10(self.time)

        b, p, SSE, R2, SEE = methods.least_squares(X, y)

        # Setup results
        self.preds = p
        self.C_avg = b[-1]
        self.C_heat = {h: b[-1] for h in self.heat_indices.keys()}
        self.polyavg = b[:-1]
        self.R2 = R2
        self.SSE = SSE
        self.SEE = SEE
        self.SEE_heat = SEE
        self.R2_heat = R2
        rms = np.sqrt(np.mean((p - y) ** 2.0))
        self.heat_rms = {h: rms for h in self.heat_indices.keys()}

        return self


class LotCenteredAnalysis(PolynomialAnalysis):
    """
    Do an uncentered analysis of the data

    Args:
        TTP:                        time-temperature parameter
        order:                      polynomial order
        data:                       dataset as a pandas dataframe

    Keyword Args:
        time_field (str):           field in array giving time, default is
                                    "Life (h)"
        temp_field (str):           field in array giving temperature, default
                                    is "Temp (C)"
        stress_field (str):         field in array giving stress, default is
                                    "Stress (MPa)"
        heat_field (str):           filed in array giving heat ID, default is
                                    "Heat/Lot ID"
        input_temp_units (str):     temperature units, default is "C"
        input_stress_units (str):   stress units, default is "MPa"
        input_time_units (str):     time units, default is "hr"
        analysis_temp_units (str):  temperature units for analysis,
                                    default is "K"
        analysis_stress_units (str):    analysis stress units, default is
                                        "MPa"
        analysis_time_units (str):  analysis time units, default is "hr"
    """

    def analyze(self):
        """
        Actually do the regression analysis and set standard properties
        """
        # Setup the lot matrix
        C = np.zeros((len(self.stress), self.nheats + 1))
        C[:, 0] = -1.0

        for i, inds in enumerate(self.heat_indices.values()):
            C[inds, i + 1] = -1.0

        # Setup the correlation matrix
        X = np.concatenate(
            (
                np.vander(np.log10(self.stress), N=self.order + 1)
                * self.TTP.stress_transform(self.time, self.temperature)[:, None],
                C,
            ),
            axis=1,
        )
        y = self.time_sign * np.log10(self.time)

        b, p, SSE, R2, SEE = methods.least_squares(X, y)

        C_avg = sum(
            (b[self.order + 1] + b[self.order + 1 + i + 1]) * len(inds)
            for i, (h, inds) in enumerate(self.heat_indices.items())
        ) / len(self.stress)

        # Now go back and calculate the SEE and the R2 values as if you have a random heat
        poly = b[: self.order + 1]
        p_prime = self.TTP.predict(poly, C_avg, self.stress, self.temperature)
        e_prime = y - p_prime
        SEE_prime = np.sqrt(np.sum(e_prime**2.0) / (X.shape[0] - self.order - 2))
        ybar = np.mean(y)
        SST = np.sum((y - ybar) ** 2.0)
        R2_heat = 1.0 - np.sum(e_prime**2.0) / SST

        # Set standard properties
        self.preds = p
        self.C_avg = C_avg
        self.C_heat = {
            h: b[self.order + 1] + b[self.order + 1 + i + 1]
            for i, h in enumerate(self.heat_indices.keys())
        }
        self.polyavg = poly
        self.R2 = R2
        self.SSE = SSE
        self.SEE = SEE
        self.SEE_heat = SEE_prime
        self.R2_heat = R2_heat
        self.heat_rms = {
            h: np.sqrt(np.mean(e_prime[inds] ** 2.0))
            for h, inds in self.heat_indices.items()
        }

        return self


class TTP(abc.ABC):
    """
    Superclass for all time-temperature parameters, currently doesn't do anything
    """

    @abc.abstractmethod
    def stress_transform(self, time, temperature):
        """
        Transform the stress

        Parameters:
            time:           time data
            temperature:    temperature data
        """
        return

    @abc.abstractmethod
    def predict(self, poly, C, stress, temperature):
        """
        Make a prediction in log time for a set of points

        Args:
            poly:           calibrated polynomial
            C:              calibrated TTP
            stress:         stress data
            temperature:    temperature data
        """
        return

    @abc.abstractmethod
    def value(self, C, time, temperature, time_sign=1.0):
        """
        Actually calculate the value of the time-temperature
        parameter

        Args:
            C:              calibrated TTP
            time:           time values
            temperature:    temperature values
        """
        return


class LarsonMillerParameter(TTP):
    """
    Larson-Miller parameters
    """

    def stress_transform(self, time, temperature):
        """
        Transform the stress

        Parameters:
            time:           time data
            temperature:    temperature data
        """
        return 1.0 / temperature

    def predict(self, poly, C, stress, temperature):
        """
        Make a prediction in log time for a set of points

        Args:
            poly:           calibrated polynomial
            C:              calibrated TTP
            stress:         stress data
            temperature:    temperature data
        """
        return np.polyval(poly, np.log10(stress)) / temperature - C

    def value(self, C, time, temperature, time_sign=1.0):
        """
        Actually calculate the value of the time-temperature
        parameter

        Args:
            C:              calibrated TTP
            time:           time values
            temperature:    temperature values
        """
        return temperature * (time_sign * np.log10(time) + C)
