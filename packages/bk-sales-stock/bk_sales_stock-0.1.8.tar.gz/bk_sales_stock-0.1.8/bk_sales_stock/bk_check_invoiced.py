from openpyxl import load_workbook

def check_daily_sales(sales_xlsx, invoiced_xlsx):
    
    wb_sales = load_workbook(sales_xlsx, data_only=True)
    ws_sales = wb_sales.active

    wb_invoiced = load_workbook(invoiced_xlsx)
    ws_invoiced = wb_invoiced.active

    key_mapping = {"LIFEBUOY SOAP TOTAL 12X6X70G":"LIFEBUOY PW BAR SOAP TOTAL 70G",
    "LIFEBUOY SOAP BAR LEMON 12X6X70G":"LIFEBUOY PW BAR LEMON FRESH 70G",
    "LIFEBUOY SOAP BAR TOTAL 6X6X150G ":"LIFEBUOY SKIN CLNSNG BAR TOTAL 150G",
    "LIFEBUOY SOAP BAR LEMON 6X6X150G ":"LIFEBUOY SKIN CLEANSING BAR LEMON 150G",
    "LIFEBUOY T 144*20G":"LIFEBUOY SKIN CLEANSING BAR TOTAL 20G",
    "KNORR_AIO":"KNORR BOUILLON CUBES BEEF 8G",
    "KNORR CUBE CHICKEN":"KNORR BOUILLONS CHICKEN CUBES 8G",
    "SIGNAL 72*60G":"SIGNAL TOTHPASTE ESS CAVITY FIGHTER 60G",
    "SIGNAL 48*140G":"SIGNAL TOTHPASTE ESS CAVITY FIGHTER 140G",
    "SUNLIGHT POWDER YELLOW 100*40G":"SUNLIGHT NM STD HW POWDER 40G",
    "SUNLIGHT POWDER 72*90G":"SUNLIGHT NM STD HW POWDER 90G",
    "SUNLIGHT BAR 50*200G":"SUNLIGHT DA HARD SOAP WHITE 200G",
    "SUNLIGHT POWDER 24*500G":"SUNLIGHT NM STD POWDER 500G",
    "SUNLIGHT POWDER 1KG":"SUNLIGHT NM STD HW POWDER 1KG",
    "SUNLIGHT 5KG":"SUNLIGHT NM STD POWDER 5KG",
    "SS COC SHA 350ML":"SUNSILK SHAMPOO COCONUT 350ML",
    "SS COC COND 350ML":"SUNSILK REG RINSE OUT COND COCONUT 350ML",
    "SS AVO SHA 350ML":"SUNSILK SHAMPOO AVOCADO 350ML",
    "SS AVO COND 350ML":"SUNSILK REG RINSE OUT COND AVOCADO 350ML",
    "SS COC SHA 700ML":"SUNSILK SHAMPOO COCONUT 700ML",
    "SS COC COND 700ML":"SUNSILK REG RINSE OUT COND COCONUT 700ML",
    "SS AVO SHA 700ML":"SUNSILK SHAMPOO AVOCADO 700ML",
    "SS AVO COND 700ML":"SUNSILK REG RINSE OUT COND AVOCADO 700ML",
    "OMO 100*40G":"OMO NM STD HW POWDER GAIA 40G",
    "OMO POWDER 72*100G":"OMO NM STD POWDER 100G",
    "OMO POWDER 24*500G":"OMO NM STD HW POWDER GAIA 500G",
    "OMO POWDER 12*1KG":"OMO NM STD HW POWDER GAIA 1KG",
    "OMO POWDER 4*3KG":"OMO NM STD HW POWDER GAIA 3KG",
    "LUX SOAP SOFT CARESS 12X6X70G":"LUX SKIN CLEANSING BAR SOFT CARESS 70G",
    "LUX SOAP SOFT TOUCH 12X6X70G":"LUX SKIN CLEANSING BAR SOFT TOUCH 70G",
    "LUX SOAP SOFT CARESS 150G":"LUX SKIN CLEANSING BAR SOFT CARESS 150G",
    "LUX SOAP SOFT TOUCH 150G":"LUX SKIN CLEANSING BAR SOFT TOUCH 150G",
    "SS COC SHA  15ML":"SUNSILK SHAMPOO COCONUT 15ML",
    "SS COC CON15ML":"SUNSILK REG RINSE OUT COND COCONUT 15ML",
    "LIFEBUOY L 144*20G":"LIFEBUOY SKN CLNG BAR LEMON FRESH 20G"}

    # Read invoiced data into a dictionary for fast lookup
    invoiced_data = {}
    for row in ws_invoiced.iter_rows(min_row=17, values_only=True):  
        inv_key = row[24]  # Column Y
        inv_value = row[40]  # Column AO
        if inv_key:
            invoiced_data[inv_key] = inv_value

    discrepancies_found = False 

    
    sales_row = 3
    while True:
        sales_key = ws_sales[f"A{sales_row}"].value  # Column A
        if not sales_key:
            break  

        sales_value = ws_sales[f"L{sales_row}"].value  # Column L


        try:
            sales_value = float(sales_value) if sales_value is not None else 0
        except ValueError:
            sales_value = 0


        mapped_key = key_mapping.get(sales_key)
        if not mapped_key:
            sales_row += 1
            continue  

        inv_value = invoiced_data.get(mapped_key)

        if inv_value is not None:

            try:
                inv_value = float(inv_value) if inv_value is not None else 0
            except ValueError:
                inv_value = 0

            if sales_value != inv_value:
                print(f"{sales_key} | Sales: {sales_value} | Invoiced: {inv_value}")
                discrepancies_found = True  

        else:
            if sales_value > 0:
                print(f"Warning: No matching invoiced record found for {sales_key}  with Sales: {sales_value}")
                discrepancies_found = True  

        sales_row += 1

    if not discrepancies_found:
        print("Good to go! All sales match the invoiced records.")

