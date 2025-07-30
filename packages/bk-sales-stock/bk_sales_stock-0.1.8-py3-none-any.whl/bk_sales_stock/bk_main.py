#beki
import os
from openpyxl import Workbook, load_workbook
from .bk_prices import rt_prices, ws_prices

home_dir = os.path.expanduser("~")
bk_main_stock = os.path.join(home_dir, "Desktop","bk","bk_main_stock.xlsx")
bk_ADD_STOCK = os.path.join(home_dir, "Desktop","bk","bk_ADD_STOCK.xlsx")
bk_rt = os.path.join(home_dir, "Desktop","bk","bk_rt.xlsx")
bk_ws = os.path.join(home_dir, "Desktop","bk","bk_ws.xlsx")


#load workbook
wb = load_workbook(bk_main_stock)
wb2 = load_workbook(bk_ADD_STOCK)
wb4 = load_workbook(bk_rt, data_only = True)
wb_rt = load_workbook(bk_rt)
wb5 = load_workbook(bk_ws, data_only = True)
wb_ws = load_workbook(bk_ws)

ws = wb['FUN']
ws2 = wb2['FUN']
add = wb2['ADD']

rt_sales = wb4['BEKI']
cls_rt_sales = wb_rt['BEKI']
rt_price_update = wb_rt['FUN']

wh_s = wb5['BEKI']
cls_wh_s = wb_ws['BEKI']
ws_price_update = wb_ws['FUN']

# add section
def add_stock():

	for i in range(2,38):
		cz = add['B'+str(i)].value
		dz = add['C'+str(i)].value
		pc = add['D'+str(i)].value
		if cz == None:
			cz = 0
		if dz == None:
			dz = 0
		if pc == None:
			pc = 0
		total_pcs = cz*ws2['C'+str(i)].value+ dz*ws2['D'+str(i)].value +pc
		ws['A'+str(i)] = ws['A'+str(i)].value + total_pcs
		
	wb.save(bk_main_stock)
	print("stock added successfully")


# deduction rt
def deduct_rt_stock():

	for i in range(2,38):
		if rt_sales['L'+str(i+1)].value ==None:
			rt_sales['L'+str(i+1)].value = 0
		ws['A'+str(i)] = ws['A'+str(i)].value - rt_sales['L'+str(i+1)].value

	wb.save(bk_main_stock)
	print("retail sales deduct from main stock successfully")

#deduction ws
def deduct_ws_stock():

	for i in range(2,38):
		if wh_s['L'+str(i+1)].value == None:
			wh_s['L'+str(i+1)].value =0
		ws['A'+str(i)] = ws['A'+str(i)].value - wh_s['L'+str(i+1)].value

	wb.save(bk_main_stock)
	print("wholesale sales deduct from main stock successfully")

def clear():
	
	for i in range(2,40):
		for j in ['B','C','D']:
			add[f'{j}{i}'] = ''
		for k in ['B','C','D','E','F','G']:
			cls_rt_sales[f'{k}{i+1}'] = ''
			cls_wh_s[f'{k}{i+1}'] = ''

	wb2.save(bk_ADD_STOCK)
	wb_rt.save(bk_rt)
	wb_ws.save(bk_ws)

def update_prices_rt():

	row = 3
	col = 9 
	for key, value in rt_prices.items():
		rt_price_update.cell(row=row, column=col, value=key)  # Product name
		rt_price_update.cell(row=row, column=col + 1, value=value)  # Product price
		row += 1
	wb_rt.save(bk_rt)
	print("retail price update successfully")

def update_prices_ws():

	row = 3
	col = 9 
	for key, value in ws_prices.items():
		ws_price_update.cell(row=row, column=col, value=key)  # Product name
		ws_price_update.cell(row=row, column=col + 1, value=value)  # Product price
		row += 1
	wb_ws.save(bk_ws)
	print("wholesale price update successfully")

def update_prices_ms():

	row = 2
	col = 9 
	for key, value in rt_prices.items():
		ws.cell(row=row, column=col, value=key)  # Product name
		ws.cell(row=row, column=col + 1, value=value)  # Product price
		row += 1
	wb.save(bk_main_stock)
	print("main stock price update successfully")




if __name__ == '__main__':
	add_stock()
	deduct_rt_stock()
	deduct_ws_stock()
	clear()
	update_prices_rt()
	update_prices_ws()
	update_prices_ms()

	wb2.close()
	wb_rt.close()
	wb_ws.close()
	wb.close()