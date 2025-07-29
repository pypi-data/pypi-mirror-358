import copy
import openpyxl
from openpyxl.utils import get_column_letter


class ManagerExcel:
    @staticmethod
    def excel_format(path, save_path):
        """
        copy excel file and keep format
        :param path: source file path
        :param save_path: save path
        :return: save path
        """
        wb = openpyxl.load_workbook(path)
        wb_new = openpyxl.Workbook()

        sheetnames = wb.sheetnames
        for sheetname in sheetnames:

            sheet = wb[sheetname]
            sheet2 = wb_new.create_sheet(sheetname)

            # copy tab color
            sheet2.sheet_properties.tabColor = sheet.sheet_properties.tabColor

            # start to process merged cells in the form of "(,)", replace "(,)"'  find the merged cells
            wm = list(sheet.merged_cells)
            if len(wm) > 0:
                for i in range(0, len(wm)):

                    cell2 = str(wm[i]).replace("(,)", "")
                    sheet2.merge_cells(cell2)

            # after traversing, write data first
            for i, row in enumerate(sheet.iter_rows()):
                sheet2.row_dimensions[i + 1].height = sheet.row_dimensions[i + 1].height
                for j, cell in enumerate(row):

                    sheet2.column_dimensions[get_column_letter(j + 1)].width = (
                        sheet.column_dimensions[get_column_letter(j + 1)].width
                    )
                    sheet2.cell(row=i + 1, column=j + 1, value=cell.value)

                    source_cell = sheet.cell(i + 1, j + 1)
                    target_cell = sheet2.cell(i + 1, j + 1)
                    target_cell.fill = copy.copy(source_cell.fill)

                    # the default style is Normal, if it is the default style, return False, do not trigger if, otherwise copy
                    if source_cell.has_style:

                        # the StyleableObject implements storing styles in a single list _style, and the style properties on the cell are actually the getter and setter of that array, so you can use the following method to clone styles faster
                        target_cell._style = copy.copy(source_cell._style)

                        # copy font size
                        target_cell.font = copy.copy(source_cell.font)

                        # copy border
                        target_cell.border = copy.copy(source_cell.border)

                        # copy fill style
                        target_cell.fill = copy.copy(source_cell.fill)

                        # copy font style
                        target_cell.number_format = copy.copy(source_cell.number_format)

                        # copy style protection
                        target_cell.protection = copy.copy(source_cell.protection)

                        # copy alignment style
                        target_cell.alignment = copy.copy(source_cell.alignment)

        if "Sheet" in wb_new.sheetnames:
            del wb_new["Sheet"]
        wb_new.save(save_path)

        wb.close()
        wb_new.close()
        return save_path
