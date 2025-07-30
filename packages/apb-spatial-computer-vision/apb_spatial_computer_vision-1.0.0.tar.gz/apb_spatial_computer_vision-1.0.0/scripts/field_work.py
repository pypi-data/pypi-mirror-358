from openpyxl import load_workbook

from apb_spatial_computer_vision import *

import geopandas as gpd , pandas as pd, numpy as np
from pyproj import Proj,Transformer,CRS

np.set_printoptions(suppress=True)
# Custom formatting function
def sci_notation_formatter(x):
    if abs(x) < 1e-6:
        return f"{x:.2e}"
    else:
        return f"{x}"


np.set_printoptions(suppress=True)

def vector_deg_to_DMS(alpha):
        alpha=np.array(alpha)
        dd=np.abs(np.astype(alpha,np.uint8))
        min=np.abs(np.astype((np.abs(alpha)-dd)*60,np.uint8))
        sec=np.abs(((((np.abs(alpha)-dd)*60)-min)*60))
        signs=np.array(np.sign(alpha))
        return ((signs*dd).astype(np.int8),(signs*min).astype(np.int8),signs*sec)
        
def quality_control_point_creation(from_crs=25831,geographic_crs=4258):
    """Data processing from the points with averaging
    """
    x_col='Coordenada X'
    y_col='Coordenada Y'
    z_col='Altura Ortom.'
    h_col='Altura Elipsoidal'
    id_col='Punto Id'
    std_x='D.Est X'
    std_y='D.Est Y'
    std_H='D.Est Altura Ortom'
    std_h='D.Est. Alt. Elip.'
    N='Ondulación Geoidal'
    date='Fecha/Hora'
    K='Factor de escala de proyección'
    df=pd.read_csv(os.path.join(DATA_DIR,'INFINITY_COMPROV_ORTO.txt'),header=0,sep=';')
    gdf=gpd.GeoDataFrame(df,geometry=gpd.points_from_xy(df[x_col],df[y_col]),crs=from_crs)
    
    projection=Proj(from_crs)
    transformer=Transformer.from_crs(from_crs,geographic_crs)

    
    gdf=gdf.to_crs(geographic_crs)
    gdf['LAT'],gdf['LON'],gdf['h']=transformer.transform(gdf[x_col],gdf[y_col],gdf[z_col])
    gdf['LAT_moved'],gdf['LON_moved'],gdf['h_moved']=transformer.transform(gdf[x_col]+gdf[std_x],gdf[y_col]+gdf[std_y],gdf[z_col])
    gdf['STD_LAT']=np.array([f"{x:.6f}" for x in (np.array(np.abs(gdf['LAT_moved']-gdf['LAT'])*3600)).round(6)])
    gdf['STD_LON']=np.array([f"{x:.6f}" for x in (np.array(np.abs(gdf['LON_moved']-gdf['LON'])*3600)).round(6)])

    ltd,ltm,lts=vector_deg_to_DMS(gdf['LAT'])
    gdf[['LTD','LTM','LTS']]=pd.DataFrame({'LTD':ltd,'LTM':ltm,'LTS':lts.round(3)})
    
    lnd,lnm,lns=vector_deg_to_DMS(gdf['LON'])
    gdf[['LND','LNM','LNS']]=pd.DataFrame({'LND':lnd,'LNM':lnm,'LNS':lns.round(3)})
    gdf[K]=gdf[K].round(7)
    #gdf['K']=projection.get_factors(gdf['LON'],gdf['LAT']).meridional_scale.round(7)
    gdf['DATA']=np.array([i.split(' ')[0] for i in gdf[date]]).astype(str)
    file_path = os.path.join(DATA_DIR,'SURVEYING.xlsx')
    
    
    workbook = load_workbook(file_path,data_only=False)
    for sheet in workbook.sheetnames:
        workbook[sheet].sheet_state = 'visible'

    sheet = workbook.active

    df_excel = pd.read_excel(file_path, sheet_name=sheet.title)
    df_excel = df_excel.reindex(range(len(gdf)), fill_value=np.nan)
    df_excel[['X_UTM','Y_UTM','LAT','LTD','LTM','LTS','LON','LND','LNM','LNS','K','H','h','N','CODI','STD_X','STD_Y','STD_H','STD_h','STD_LAT','STD_LON','DATA']] = gdf[[x_col,y_col,'LAT','LTD','LTM','LTS','LON','LND','LNM','LNS',K,z_col,h_col,N,id_col,std_x,std_y,std_H,std_h,'STD_LAT','STD_LON','DATA']].values
    df_excel['CRS']='ETRS89'
    df_excel['GEOIDE']='EGM08D595'
    df[[x_col,y_col,z_col,std_x,std_y]].round(3)
    df_excel=df_excel.astype(str)
    df_excel.to_excel(file_path,engine='openpyxl', sheet_name=sheet.title, index=False)
    workbook.close()
    df_excel.to_csv(os.path.join(OUT_DIR,'ORTOFOTO_CHECK.CSV'))
    gdf.to_file(os.path.join(DATA_DIR,'ORTOFOTO_CHECK.geojson'))
quality_control_point_creation()