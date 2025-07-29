# -*- coding:utf-8 -*-
from csv import reader as csv_reader, writer as csv_writer
from pathlib import Path
from time import sleep

from openpyxl.reader.excel import load_workbook

from .base import BaseRecorder
from .cell_style import NoneStyle
from .setter import RecorderSetter, set_csv_header
from .tools import (ok_list_str, process_content_xlsx, process_content_json, get_key_cols,
                    get_csv, parse_coord, do_nothing, Header, get_wb, get_ws,
                    get_real_coord, is_sigal_data, is_1D_data, get_real_col, data2ws, styles2ws, get_real_row)


class Recorder(BaseRecorder):
    _SUPPORTS = ('csv', 'xlsx', 'txt', 'jsonl', 'json')

    def __init__(self, path=None, cache_size=1000):
        self._header = {None: None}
        self._methods = {'xlsx': self._to_xlsx_fast,
                         'csv': self._to_csv_fast,
                         'txt': self._to_txt_fast,
                         'jsonl': self._to_jsonl_fast,
                         'json': self._to_json_fast,
                         'addData': not_type,
                         # fast模式下非添加data使用的方法
                         'setImg': do_nothing,
                         'setLink': do_nothing,
                         'setStyle': do_nothing,
                         'setHeight': do_nothing,
                         'setWidth': do_nothing,
                         # 写入不同类型数据时使用的方法
                         'img': img2ws,
                         'link': link2ws,
                         'style': styles2ws,
                         'height': height2ws,
                         'width': width2ws,
                         'data': data2ws}
        super().__init__(path=path, cache_size=cache_size)
        self._data = {}
        self._delimiter = ','  # csv文件分隔符
        self._quote_char = '"'  # csv文件引用符
        self._follow_styles = False
        self._row_height = None
        self._styles = None
        self._header_row = {None: 1}
        self._fast = True
        self._link_style = None
        self.data_col = 1

    def _set_methods(self, file_type):
        self._methods[file_type] = getattr(self, f'_to_{file_type}_fast')
        if file_type == 'xlsx':
            self._methods['setImg'] = self._set_img
            self._methods['setLink'] = self._set_link
            self._methods['setStyle'] = self._set_styles
            self._methods['setHeight'] = self._set_row_height
            self._methods['setWidth'] = self._set_col_width
            self._methods['addData'] = self._add_data_any
        else:
            self._methods['setImg'] = do_nothing
            self._methods['setLink'] = do_nothing
            self._methods['setStyle'] = do_nothing
            self._methods['setHeight'] = do_nothing
            self._methods['setWidth'] = do_nothing
            self._methods['addData'] = self._add_data_txt

    @property
    def set(self):
        if self._setter is None:
            self._setter = RecorderSetter(self)
        return self._setter

    @property
    def delimiter(self):
        return self._delimiter

    @property
    def quote_char(self):
        return self._quote_char

    @property
    def header(self):
        if self.type not in ('csv', 'xlsx'):
            raise TypeError('header属性只支持csv和xlsx类型文件。')
        return get_header(self)

    def add_data(self, data, coord=None, table=None):
        coord = parse_coord(coord, self.data_col)
        data, data_len = self._handle_data(data, coord)
        self._add(data, table,
                  True if self._fast and coord[0] else False,  # xlsx不需要fast模式
                  data_len, self._methods['addData'])

    def _handle_data(self, data, coord):
        if is_sigal_data(data):
            data = {'type': 'data', 'data': [self._make_final_data(self, (data,))], 'coord': coord}
            data_len = 1
        elif not data:
            data = {'type': 'data', 'data': [self._make_final_data(self, tuple())], 'coord': coord}
            data_len = 1
        elif is_1D_data(data):
            data = {'type': 'data', 'data': [self._make_final_data(self, data)], 'coord': coord}
            data_len = 1
        else:  # 二维数组
            data = {'type': 'data', 'coord': coord,
                    'data': [self._make_final_data(self, (d,)) if is_sigal_data(d)
                             else self._make_final_data(self, d) for d in data]}
            data_len = len(data)
        return data, data_len

    def _add(self, data, table, to_slow, num, add_method):
        while self._pause_add:  # 等待其它线程写入结束
            sleep(.02)

        if to_slow:
            self._slow_mode()

        if table is None:
            table = self._table
        elif table is True:
            table = None

        add_method(data, table)

        self._data_count += num
        if 0 < self.cache_size <= self._data_count:
            self.record()

    def _add_data_any(self, data, table):
        if (self._data.get(table, 0) != 0 and self._data[table]
                and data['coord'] == self._data[table][-1]['coord']
                and self._data[table][-1]['type'] == 'data'):
            self._data[table][-1]['data'].extend(data['data'])
        else:
            self._data.setdefault(table, []).append(data)

    def _add_data_txt(self, data, table):
        if self._data.get(None, None) and data['coord'] == self._data[None][-1]['coord']:
            self._data[table][-1]['data'].extend(data['data'])
        else:
            self._data.setdefault(None, []).append(data)

    def _add_others(self, data, table):
        self._data.setdefault(table, []).append(data)

    def _set_link(self, coord, link, content=None, table=None):
        self._add({'type': 'link', 'link': link, 'content': content,
                   'coord': parse_coord(coord, self.data_col)}, table, self._fast, 1, self._add_others)

    def _set_img(self, coord, img_path, width=None, height=None, table=None):
        self._add({'type': 'img', 'imgPath': img_path, 'width': width, 'height': height,
                   'coord': parse_coord(coord, self.data_col)}, table, self._fast, 1, self._add_others)

    def _set_styles(self, coord, styles, replace=True, table=None):
        if isinstance(coord, str):
            if ':' in coord:
                real, coord = coord, (1, 1)
            elif coord.isdigit() or (coord[0] == '-' and coord[1:].isdigit()):
                real, coord = int(coord), (1, 1)
            else:
                real, coord = None, parse_coord(coord, self.data_col)
        elif isinstance(coord, int):
            real, coord = coord, (1, 1)
        else:
            real, coord = None, parse_coord(coord, self.data_col)
        self._add({'type': 'style', 'mode': 'replace' if replace else 'cover', 'real_coord': real,
                   'styles': styles, 'coord': coord}, table, self._fast, 1, self._add_others)

    def _set_row_height(self, row, height, table=None):
        if not row:
            raise ValueError('row不能为0或None。')
        self._add({'type': 'height', 'row': row, 'height': height}, table, self._fast, 1, self._add_others)

    def _set_col_width(self, col, width, table=None):
        if not col:
            raise ValueError('col不能为0或None。')
        self._add({'type': 'width', 'col': col, 'width': width}, table, self._fast, 1, self._add_others)

    def set_link(self, coord, link, content=None, table=None):
        self._methods['setLink'](coord, link, content, table)

    def set_img(self, coord, img_path, width=None, height=None, table=None):
        self._methods['setImg'](coord, img_path, width, height, table)

    def set_styles(self, coord, styles, replace=True, table=None):
        self._methods['setStyle'](coord, styles, replace, table)

    def set_row_height(self, row, height, table=None):
        self._methods['setHeight'](row, height, table)

    def set_col_width(self, col, width, table=None):
        self._methods['setWidth'](col, width, table)

    def rows(self, key_cols=True, sign_col=True, is_header=False,
             signs=None, deny_sign=False, count=None, begin_row=None):
        if not self._path or not Path(self._path).exists():
            raise RuntimeError('未指定文件路径或文件不存在。')
        if self.type == 'xlsx':
            wb = load_workbook(self.path, data_only=True, read_only=True)
            if self.table and self.table not in [i.title for i in wb.worksheets]:
                raise RuntimeError(f'xlsx文件未包含指定工作表：{self.table}')
            ws = wb[self.table] if self.table else wb.active
            if ws.max_column is None:  # 遇到过read_only时无法获取列数的文件
                wb.close()
                wb = load_workbook(self.path, data_only=True)
                ws = wb[self.table] if self.table else wb.active

            method = get_xlsx_rows

        elif self.type == 'csv':
            ws = None
            method = get_csv_rows
        else:
            raise RuntimeError('rows()方法只支持xlsx和csv格式。')

        header = get_header(self, ws)
        if not begin_row:
            begin_row = self._header_row.get(self.table, self._header_row[None]) + 1

        if sign_col is not True:
            sign_col = header.get_num(sign_col, is_header=is_header) or 1
        if not isinstance(signs, (list, tuple, set)):
            signs = (signs,)
        key_cols = get_key_cols(key_cols, header, is_header)

        return method(self, header=header, key_cols=key_cols, begin_row=begin_row, sign_col=sign_col,
                      sign=signs, deny_sign=deny_sign, count=count, ws=ws)

    def _record(self):
        self._methods[self.type]()
        if not self._fast:
            self._fast_mode()

    def _fast_mode(self):
        self._methods['csv'] = self._to_csv_fast
        self._methods['txt'] = self._to_txt_fast
        self._methods['json'] = self._to_json_fast
        self._methods['jsonl'] = self._to_jsonl_fast
        self._fast = True

    def _slow_mode(self):
        self._methods['csv'] = self._to_csv_slow
        self._methods['txt'] = self._to_txt_slow
        self._methods['json'] = self._to_json_slow
        self._methods['jsonl'] = self._to_jsonl_slow
        self._fast = False

    def _to_xlsx_fast(self):
        wb, new_file = get_wb(self)
        tables = wb.sheetnames
        rewrite_method = 'make_num_dict_rewrite' if self._auto_new_header else 'make_num_dict'

        for table, data in self._data.items():
            ws, new_sheet = get_ws(wb, table, tables, new_file)
            new_file = False
            if table is None:
                if ws.title not in self._header:
                    self._header[ws.title] = self._header[None]
                if ws.title not in self._header_row:
                    self._header_row[ws.title] = self._header_row[None]

            begin_row = True
            if new_sheet:
                begin_row = handle_new_sheet(self, ws, data)
            elif self._header.get(ws.title, None) is None:
                self._header[ws.title] = Header([c.value for c in ws[self._header_row[ws.title]]])

            header = self._header[ws.title]
            rewrite = False
            if not begin_row and not data[0]['coord'][0]:
                cur = data[0]
                rewrite = self._methods[cur['type']](
                    **{'recorder': self,
                       'ws': ws,
                       'data': cur,
                       'coord': (1, get_real_col(cur.get('coord', (1, 1)), ws.max_column)),
                       'new_row': not cur.get('coord', (1, 1))[0],
                       'header': header,
                       'rewrite': rewrite,
                       'rewrite_method': rewrite_method})
                data = data[1:]

            for cur in data:
                rewrite = self._methods[cur['type']](
                    **{'recorder': self,
                       'ws': ws,
                       'data': cur,
                       'coord': get_real_coord(cur.get('coord', (1, 1)), ws.max_row, ws.max_column),
                       'new_row': not cur.get('coord', (1, 1))[0],
                       'header': header,
                       'rewrite': rewrite,
                       'rewrite_method': rewrite_method})

            if rewrite:
                for c in range(1, ws.max_column + 1):
                    ws.cell(self._header_row[ws.title], c, value=header[c])

        wb.save(self.path)
        wb.close()

    def _to_csv_fast(self):
        file, new_csv = get_csv(self)
        writer = csv_writer(file, delimiter=self.delimiter, quotechar=self.quote_char)
        get_and_set_csv_header(self, new_csv, file, writer)
        rewrite_method = 'make_insert_list_rewrite' if self._auto_new_header else 'make_insert_list'

        rewrite = False
        header = self._header[None]
        for d in self._data[None]:
            col = 1 if d['coord'][1] == 1 else get_real_col(d['coord'][1],
                                                            len(header) if self._header_row[None] > 0 else 1)
            for data in d['data']:
                data, rewrite = header.__getattribute__(rewrite_method)(data, 'csv', rewrite)
                data = [None] * (col - 1) + data
                writer.writerow(data)
        file.close()

        if rewrite:
            set_csv_header(self, self._header[None], self._header_row[None])

    def _to_csv_slow(self):
        file, new_csv = get_csv(self)
        writer = csv_writer(file, delimiter=self.delimiter, quotechar=self.quote_char)
        get_and_set_csv_header(self, new_csv, file, writer)
        file.seek(0)
        reader = csv_reader(file, delimiter=self.delimiter, quotechar=self.quote_char)
        lines = list(reader)
        lines_count = len(lines)
        header = self._header[None]

        rewrite = False
        method = 'make_change_list_rewrite' if self._auto_new_header else 'make_change_list'
        for i in self._data[None]:
            data = i['data']
            row, col = get_real_coord(i['coord'], lines_count, len(header) if self._header_row[None] > 0 else 1)
            for r, da in enumerate(data, row):
                add_rows = r - lines_count
                if add_rows > 0:  # 若行数不够，填充行数
                    [lines.append([]) for _ in range(add_rows)]
                    lines_count += add_rows
                row_num = r - 1
                lines[row_num], rewrite = self._header[None].__getattribute__(method)(lines[row_num], da, col,
                                                                                      'csv', rewrite)

        if rewrite:
            [lines.append([]) for _ in range(self._header_row[None] - lines_count)]  # 若行数不够，填充行数
            lines[self._header_row[None] - 1] = list(header.num_key.values())

        file.close()
        writer = csv_writer(open(self.path, 'w', encoding=self.encoding, newline=''),
                            delimiter=self.delimiter, quotechar=self.quote_char)
        writer.writerows(lines)

    def _to_txt_fast(self):
        with open(self.path, 'a+', encoding=self.encoding) as f:
            all_data = []
            for data in self._data[None]:
                for d in data['data']:
                    all_data.append(' '.join(ok_list_str(d)))
            f.write('\n'.join(all_data) + '\n')

    def _to_jsonl_fast(self):
        from json import dumps
        with open(self.path, 'a+', encoding=self.encoding) as f:
            all_data = []
            for data in self._data[None]:
                for d in data['data']:
                    all_data.append(d if isinstance(d, str) else dumps(d))
            f.write('\n'.join(all_data) + '\n')

    def _to_json_fast(self):
        from json import load, dump
        if self._file_exists or Path(self.path).exists():
            with open(self.path, 'r', encoding=self.encoding) as f:
                json_data = load(f)
        else:
            json_data = []

        for i in self._data[None]:
            for data in i['data']:
                if isinstance(data, dict):
                    for k, d in data.items():
                        data[k] = process_content_json(d)
                    json_data.append(data)
                else:
                    json_data.append([process_content_json(d) for d in i])

        with open(self.path, 'w', encoding=self.encoding) as f:
            dump(json_data, f)

    def _to_txt_slow(self):
        if not self._file_exists and not Path(self.path).exists():
            with open(self.path, 'w', encoding=self.encoding):
                pass
        with open(self.path, 'r+', encoding=self.encoding) as f:
            lines = f.readlines()
            handle_txt_lines(self._data[None], lines, '\n', handle_jsonl_data)
            f.seek(0)
            f.writelines(lines)

    def _to_jsonl_slow(self):
        if not self._file_exists and not Path(self.path).exists():
            with open(self.path, 'w', encoding=self.encoding):
                pass
        with open(self.path, 'r+', encoding=self.encoding) as f:
            lines = f.readlines()
            handle_txt_lines(self._data[None], lines, '[]\n', handle_jsonl_data)
            f.seek(0)
            f.writelines(lines)

    def _to_json_slow(self):
        from json import load, dump
        if self._file_exists or Path(self.path).exists():
            with open(self.path, 'r', encoding=self.encoding) as f:
                lines = load(f)
        else:
            lines = []
        handle_txt_lines(self._data[None], lines, None, handle_json_data)
        with open(self.path, 'w', encoding=self.encoding) as f:
            dump(lines, f)


def handle_txt_lines(data_lst, lines, val, method):
    lines_len = len(lines)
    for data in data_lst:
        num = get_real_row(data['coord'][0], lines_len)
        data_end = num + len(data['data'])
        if lines_len < data_end:
            diff = data_end - lines_len - 1
            [lines.append(val) for _ in range(diff)]
            lines_len += diff
        for num, i in enumerate(data['data'], num - 1):
            method(lines, num, i)


def handle_txt_data(lines, num, data):
    lines[num] = ' '.join(ok_list_str(data)) + '\n'


def handle_jsonl_data(lines, num, data):
    from json import dumps
    lines[num] = data if isinstance(data, str) else dumps(data) + '\n'


def handle_json_data(lines, num, data):
    if isinstance(data, dict):
        for k, d in data.items():
            data[k] = process_content_json(d)
        lines[num] = data
    else:
        lines[num] = [process_content_json(d) for d in data]


def get_header(recorder, ws=None):
    header = recorder._header.get(recorder._table, None)
    if header is not None:
        return header
    if not recorder.path or not Path(recorder.path).exists():
        return None

    if recorder.type == 'xlsx':
        if not ws:
            wb = load_workbook(recorder.path)
            if not recorder.table:
                ws = wb.active
            elif recorder.table not in wb.sheetnames:
                wb.close()
                return Header()
            else:
                ws = wb[recorder.table]
        header_row = recorder._header_row.get(recorder.table, recorder._header_row[None])
        if header_row > ws.max_row:
            recorder._header[recorder.table] = Header()
        else:
            recorder._header[recorder.table] = Header(
                [i.value for i in ws[recorder._header_row.get(recorder.table, recorder._header_row[None])]])

        if not ws:
            wb.close()
        return recorder._header[recorder.table]

    if recorder.type == 'csv':
        from csv import reader
        with open(recorder.path, 'r', newline='', encoding=recorder.encoding) as f:
            u = reader(f, delimiter=recorder.delimiter, quotechar=recorder.quote_char)
            try:
                for _ in range(recorder._header_row[None]):
                    header = next(u)
            except StopIteration:  # 文件是空的
                header = []
        recorder._header[None] = Header(header)
        return recorder._header[None]


def handle_new_sheet(recorder, ws, data):
    if not recorder._header_row:
        return 0

    if recorder._header.get(ws.title, None) is not None:
        for c, h in recorder._header[ws.title].items():
            ws.cell(row=recorder._header_row[ws.title], column=c, value=h)
        begin_row = recorder._header_row

    else:
        data = get_first_dict(data)
        if data:
            header = Header([h for h in data.keys() if isinstance(h, str)])
            recorder._header[ws.title] = header
            for c, h in header.items():
                ws.cell(row=recorder._header_row[ws.title], column=c, value=h)
            begin_row = recorder._header_row

        else:
            recorder._header[ws.title] = Header()
            begin_row = 0

    return begin_row


def get_first_dict(data):
    if not data:
        return False
    elif data[0]['type'] == 'data' and data[0]['data'] and isinstance(data[0]['data'][0], dict):
        return data[0]['data'][0]


def get_xlsx_rows(recorder, header, key_cols, begin_row, sign_col, sign, deny_sign, count, ws):
    rows = ws.rows
    try:
        for _ in range(begin_row - 1):
            next(rows)
    except StopIteration:
        return []

    if sign_col is True or sign_col > ws.max_column:  # 获取所有行
        if count:
            rows = list(rows)[:count]
        if key_cols is True:  # 获取整行
            res = [header.make_row_data(ind, {col: cell.value for col, cell in enumerate(row, 1)})
                   for ind, row in enumerate(rows, begin_row)]
        else:  # 只获取对应的列
            res = [header.make_row_data(ind, {col: row[col - 1].value for col in key_cols})
                   for ind, row in enumerate(rows, begin_row)]

    else:  # 获取符合条件的行
        if count:
            res = get_xlsx_rows_with_count(key_cols, deny_sign, header, rows,
                                           begin_row, sign_col, sign, count)
        else:
            res = get_xlsx_rows_without_count(key_cols, deny_sign, header, rows, begin_row,
                                              sign_col, sign)

    ws.parent.close()
    return res


def get_csv_rows(recorder, header, key_cols, begin_row, sign_col, sign, deny_sign, count, ws):
    sign = ['' if i is None else str(i) for i in sign]
    begin_row -= 1
    res = []
    with open(recorder.path, 'r', encoding=recorder.encoding) as f:
        reader = csv_reader(f, delimiter=recorder.delimiter, quotechar=recorder.quote_char)
        lines = list(reader)
        if not lines:
            return res

        if sign_col is True:  # 获取所有行
            header_len = len(header)
            for ind, line in enumerate(lines[begin_row:count + 1 if count else None], begin_row + 1):
                if key_cols is True:  # 获取整行
                    if not line:
                        res.append(header.make_row_data(ind, {col: '' for col in range(1, header_len + 1)}))
                    else:
                        line_len = len(line)
                        x = max(header_len, line_len)
                        res.append(header.make_row_data(ind, {col: line[col - 1] if col <= line_len else ''
                                                              for col in range(1, x + 1)}))
                else:  # 只获取对应的列
                    x = len(line) + 1
                    res.append(header.make_row_data(ind, {col: line[col - 1] if col < x else '' for col in key_cols}))

        else:  # 获取符合条件的行
            sign_col -= 1
            if count:
                get_csv_rows_with_count(lines, begin_row, sign_col, sign, deny_sign,
                                        key_cols, res, header, count)
            else:
                get_csv_rows_without_count(lines, begin_row, sign_col, sign, deny_sign,
                                           key_cols, res, header)

    return res


def get_xlsx_rows_with_count(key_cols, deny_sign, header, rows, begin_row, sign_col, sign, count):
    got = 0
    res = []
    if key_cols is True:  # 获取整行
        if deny_sign:
            for ind, row in enumerate(rows, begin_row):
                if got == count:
                    break
                if row[sign_col - 1].value not in sign:
                    res.append(header.make_row_data(ind, {col: cell.value for col, cell in enumerate(row, 1)}))
                    got += 1
        else:
            for ind, row in enumerate(rows, begin_row):
                if got == count:
                    break
                if row[sign_col - 1].value in sign:
                    res.append(header.make_row_data(ind, {col: cell.value for col, cell in enumerate(row, 1)}))
                    got += 1

    else:  # 只获取对应的列
        if deny_sign:
            for ind, row in enumerate(rows, begin_row):
                if got == count:
                    break
                if row[sign_col - 1].value not in sign:
                    res.append(header.make_row_data(ind, {col: row[col - 1].value for col in key_cols}))
                    got += 1
        else:
            for ind, row in enumerate(rows, begin_row):
                if got == count:
                    break
                if row[sign_col - 1].value in sign:
                    res.append(header.make_row_data(ind, {col: row[col - 1].value for col in key_cols}))
                    got += 1
    return res


def get_xlsx_rows_without_count(key_cols, deny_sign, header, rows, begin_row, sign_col, sign):
    if key_cols is True:  # 获取整行
        if deny_sign:
            return [header.make_row_data(ind, {col: cell.value for col, cell in enumerate(row, 1)})
                    for ind, row in enumerate(rows, begin_row)
                    if row[sign_col - 1].value not in sign]
        else:
            return [header.make_row_data(ind, {col: cell.value for col, cell in enumerate(row, 1)})
                    for ind, row in enumerate(rows, begin_row)
                    if row[sign_col - 1].value in sign]

    else:  # 只获取对应的列
        if deny_sign:
            return [header.make_row_data(ind, {col: row[col - 1].value for col in key_cols})
                    for ind, row in enumerate(rows, begin_row)
                    if row[sign_col - 1].value not in sign]
        else:
            return [header.make_row_data(ind, {col: row[col - 1].value for col in key_cols})
                    for ind, row in enumerate(rows, begin_row)
                    if row[sign_col - 1].value in sign]


def get_csv_rows_with_count(lines, begin_row, sign_col, sign, deny_sign, key_cols, res, header, count):
    got = 0
    header_len = len(header)
    for ind, line in enumerate(lines[begin_row:], begin_row + 1):
        if got == count:
            break
        row_sign = '' if sign_col > len(line) - 1 else line[sign_col]
        if (row_sign not in sign) if deny_sign else (row_sign in sign):
            if key_cols is True:  # 获取整行
                if not line:
                    res.append(header.make_row_data(ind, {col: '' for col in range(1, header_len + 1)}))
                else:
                    line_len = len(line)
                    x = max(header_len, line_len)
                    res.append(header.make_row_data(ind, {col: line[col - 1] if col <= line_len else ''
                                                          for col in range(1, x + 1)}))
            else:  # 只获取对应的列
                x = len(line) + 1
                res.append(header.make_row_data(ind, {col: line[col - 1] if col < x else '' for col in key_cols}))
            got += 1


def get_csv_rows_without_count(lines, begin_row, sign_col, sign, deny_sign, key_cols, res, header):
    header_len = len(header)
    for ind, line in enumerate(lines[begin_row:], begin_row + 1):
        row_sign = '' if sign_col > len(line) - 1 else line[sign_col]
        if (row_sign not in sign) if deny_sign else (row_sign in sign):
            if key_cols is True:  # 获取整行
                if not line:
                    res.append(header.make_row_data(ind, {col: '' for col in range(1, header_len + 1)}))
                else:
                    line_len = len(line)
                    x = max(header_len, line_len)
                    res.append(header.make_row_data(ind, {col: line[col - 1] if col <= line_len else ''
                                                          for col in range(1, x + 1)}))
            else:  # 只获取对应的列
                x = len(line) + 1
                res.append(header.make_row_data(ind, {col: line[col - 1] if col < x else '' for col in key_cols}))


def get_and_set_csv_header(recorder, new_csv, file, writer):
    if not recorder._header_row:
        return

    if new_csv:
        if recorder._header[None]:
            for _ in range(recorder._header_row[None] - 1):
                writer.writerow([])
            writer.writerow(ok_list_str(recorder._header[None]))

        if recorder._header[None] is None and recorder._data_count:
            data = get_first_dict(recorder._data[None])
            if data:
                recorder._header[None] = Header([h for h in data.keys() if isinstance(h, str)])
            else:
                recorder._header[None] = Header()
            writer.writerow(ok_list_str(recorder._header[None]))
        else:
            recorder._header[None] = Header()

    elif recorder._header[None] is None:  # 从文件读取表头
        file.seek(0)
        reader = csv_reader(file, delimiter=recorder.delimiter, quotechar=recorder.quote_char)
        header = []
        try:
            for _ in range(recorder._header_row[None]):
                header = next(reader)
        except StopIteration:
            pass
        recorder._header[None] = Header(header)
        file.seek(2)


def link2ws(**kwargs):
    recorder = kwargs['recorder']
    data = kwargs['data']
    cell = kwargs['ws'].cell(*kwargs['coord'])
    has_link = bool(cell.hyperlink)
    cell.hyperlink = data['link']
    if data['content'] is not None:
        cell.value = process_content_xlsx(data['content'])
    if data['link']:
        if recorder._link_style:
            recorder._link_style.to_cell(cell, replace=False)
    elif has_link:
        NoneStyle().to_cell(cell, replace=False)


def img2ws(**kwargs):
    row, col = kwargs['coord']
    data = kwargs['data']
    ws = kwargs['ws']
    from openpyxl.drawing.image import Image
    img = Image(data['imgPath'])
    width, height = data['width'], data['height']
    if width and height:
        img.width = width
        img.height = height
    elif width:
        img.height = int(img.height * (width / img.width))
        img.width = width
    elif height:
        img.width = int(img.width * (height / img.height))
        img.height = height
    # ws.add_image(img, (row, Header._NUM_KEY[col]))
    ws.add_image(img, f'{Header._NUM_KEY[col]}{row}')


def width2ws(**kwargs):
    col, width = kwargs['data']['col'], kwargs['data']['width']
    ws = kwargs['ws']
    if isinstance(col, int):
        col = Header._NUM_KEY[col]
    for c in col.split(':'):
        if c.isdigit():
            c = Header._NUM_KEY[int(c)]
        ws.column_dimensions[c].width = width


def height2ws(**kwargs):
    row, height = kwargs['data']['row'], kwargs['data']['height']
    ws = kwargs['ws']
    if isinstance(row, int):
        ws.row_dimensions[row].height = height
    elif isinstance(row, str):
        for r in row.split(':'):
            ws.row_dimensions[int(r)].height = height


def not_type(*keys):
    raise RuntimeError('添加数据前请先指定文件路径或类型。')
